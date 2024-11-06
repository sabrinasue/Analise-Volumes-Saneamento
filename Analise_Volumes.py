import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import streamlit as st

# Configurações da página
st.set_page_config(
    layout="wide",
    page_title="Análise de Volumes"
)

# Side bar
st.sidebar.image("images/GEE.png", caption="Data Analytics")
# Markdown
st.sidebar.markdown('Desenvolvido por [Sabrina Bilio](sabrina.bilio@aegea.com.br)')

# Título da Página
st.header('Volumes por Unidades de Negócio', divider='blue')

# Legenda das cores
legenda_html = """
<table style="float: right; margin-top: 10px;">
  <tr>
    <td style="background-color: #90EE90; width: 20px; height: 20px;"></td>
    <td>Valores igual a 0</td>
  </tr>
  <tr>
    <td style="background-color: #87CEEB; width: 20px; height: 20px;"></td>
    <td>Valores duplicados</td>
  </tr>
  <tr>
    <td style="background-color: red; width: 20px; height: 20px;"></td>
    <td>50%</td>
  </tr>
  <tr>
    <td style="background-color: orange; width: 20px; height: 20px;"></td>
    <td>30%</td>
  </tr>
  <tr>
    <td style="background-color: yellow; width: 20px; height: 20px;"></td>
    <td>15%</td>
  </tr>
  <tr>
    <td style="background-color: green; width: 20px; height: 20px;"></td>
    <td>5%</td>
  </tr>
</table>
"""
st.write(legenda_html, unsafe_allow_html=True)

# Carregar arquivo Excel
uploaded_file = st.file_uploader("Selecione o arquivo Excel", type=["xlsx"])

if uploaded_file is not None:
    # Ler arquivo Excel
    df = pd.read_excel(uploaded_file)

# Preenchendo colunas mescladas
df['Unidade'] = df['Unidade'].fillna(method='ffill')
df['Grupo'] = df['Grupo'].fillna(method='ffill')

# Definindo a função para converter valores para float e formatar como strings com ponto e vírgula
def convert_and_format(value):
    try:
        float_value = float(value)
        formatted_value = '{:,.2f}'.format(float_value).replace('.', '#').replace(',', '.').replace('#', ',')
        return formatted_value
    except ValueError:
        return value
    
# Converter valores das colunas para float e formatar como strings com ponto e vírgula
for col in range(3, 9):  # Colunas correspondentes a 01/2024, 02/2024, 03/2024, 04/2024, 05/2024, 06/2024
    df.iloc[:, col] = df.iloc[:, col].apply(convert_and_format)

# Definindo a função para estilizar colunas
def estilizar_colunas(row):
    subgrupo = row['Subgrupo']
    valores = row.iloc[3:9].tolist()
    duplicados = {x for x in valores if valores.count(x) > 1 and pd.notna(x)}
    styles = ['' for _ in range(len(row))]
    
    if subgrupo in ['Captado', 'Produzido', 'Coletado', 'Tratado']:
        for col in range(3, 9):
            valor_atual = float(row[col].replace('.', '').replace(',', '.'))
            if valor_atual == 0:
                styles[col] = 'background-color: #90EE90'  # Verde claro para valores igual a 0
            elif row[col] in duplicados:
                styles[col] = 'background-color: #87CEEB' #Azul claro para valores duplicados 
            elif col > 5:  # Verificar se tem pelo menos 3 meses anteriores
                valores_anteriores = [float(val.replace('.', '').replace(',', '.')) for val in row[col-3:col] if pd.notna(val)]
                if len(valores_anteriores) >= 3:  # Verificar se tem pelo menos 3 meses anteriores
                    media_anteriores = sum(valores_anteriores) / len(valores_anteriores)
                    if media_anteriores != 0:  # Verificar se a média é diferente de zero
                        diferenca_porcentagem = ((valor_atual - media_anteriores) / media_anteriores) * 100
                        if abs(diferenca_porcentagem) >= 50:
                            styles[col] = 'background-color: red'  # Subiu ou caiu mais de 50%
                        elif abs(diferenca_porcentagem) >= 30:
                            styles[col] = 'background-color: orange'  # Subiu ou caiu mais de 30%
                        elif abs(diferenca_porcentagem) >= 15:
                            styles[col] = 'background-color: yellow'  # Subiu ou caiu mais de 15%
                        elif abs(diferenca_porcentagem) >= 5:
                            styles[col] = 'background-color: green'  # Subiu ou caiu mais de 5%
    
    return styles

def verificar_cor(row, cor):
    subgrupo = row['Subgrupo']
    valores = row.iloc[3:9].tolist()
    duplicados = {x for x in valores if valores.count(x) > 1 and pd.notna(x)}
    for col in range(3, 9):
        valor_atual = float(row[col].replace('.', '').replace(',', '.'))
        if valor_atual == 0:
            if cor == '#90EE90':
                return True
        elif row[col] in duplicados:
            if cor == '#87CEEB':
                return True
        elif col > 5:  # Verificar se tem pelo menos 3 mês anterior
            valores_anteriores = [float(val.replace('.', '').replace(',', '.')) for val in row[col-3:col] if pd.notna(val)]
            if len(valores_anteriores) >= 3:  # Verificar se tem pelo menos 3 meses anteriores
                media_anteriores = sum(valores_anteriores) / len(valores_anteriores)
                if media_anteriores != 0:  # Verificar se a média é diferente de zero
                    diferenca_porcentagem = ((valor_atual - media_anteriores) / media_anteriores) * 100
                    if cor == 'red' and (diferenca_porcentagem >= 50 or diferenca_porcentagem <= -50):
                        return True
                    elif cor == 'orange' and ((diferenca_porcentagem >= 30 and diferenca_porcentagem < 50) or (diferenca_porcentagem <= -30 and diferenca_porcentagem > -50)):
                        return True
                    elif cor == 'yellow' and ((diferenca_porcentagem >= 15 and diferenca_porcentagem < 30) or (diferenca_porcentagem <= -15 and diferenca_porcentagem > -30)):
                        return True
                    elif cor == 'green' and ((diferenca_porcentagem >= 5 and diferenca_porcentagem < 15) or (diferenca_porcentagem <= -5 and diferenca_porcentagem > -15)):
                        return True
    return False

def aplicar_estilo_cor(row, cor):
    styles = ['' for _ in range(len(row))]
    for col in range(3, 9):
        valor_atual = float(row[col].replace('.', '').replace(',', '.'))
        if cor == '#90EE90' and valor_atual == 0:
            styles[col] = f'background-color: {cor}'
        elif cor == '#87CEEB' and row[col] in {x for x in row.iloc[3:9].tolist() if row.iloc[3:9].tolist().count(x) > 1 and pd.notna(x)}:
            styles[col] = f'background-color: {cor}'
        elif col > 5:  # Verificar se tem pelo menos 3 mês anterior
            valores_anteriores = [float(val.replace('.', '').replace(',', '.')) for val in row[col-3:col] if pd.notna(val)]
            if len(valores_anteriores) >= 3:  # Verificar se tem pelo menos 3 meses anteriores
                media_anteriores = sum(valores_anteriores) / len(valores_anteriores)
                if media_anteriores != 0:  # Verificar se a média é diferente de zero
                    diferenca_porcentagem = ((valor_atual - media_anteriores) / media_anteriores) * 100
                    if cor == 'red' and (diferenca_porcentagem >= 50 or diferenca_porcentagem <= -50):
                        styles[col] = f'background-color: {cor}'
                    elif cor == 'orange' and ((diferenca_porcentagem >= 30 and diferenca_porcentagem < 50) or (diferenca_porcentagem <= -30 and diferenca_porcentagem > -50)):
                        styles[col] = f'background-color: {cor}'
                    elif cor == 'yellow' and ((diferenca_porcentagem >= 15 and diferenca_porcentagem < 30) or (diferenca_porcentagem <= -15 and diferenca_porcentagem > -30)):
                        styles[col] = f'background-color: {cor}'
                    elif cor == 'green' and ((diferenca_porcentagem >= 5 and diferenca_porcentagem < 15) or (diferenca_porcentagem <= -5 and diferenca_porcentagem > -15)):
                        styles[col] = f'background-color: {cor}'
    return styles

# Definindo as cores disponíveis
cores = {
    'Todas': 'Todas',
    '#90EE90': 'Verde Claro',
    '#87CEEB': 'Azul Claro',
    'red': 'Vermelho',
    'orange': 'Laranja',
    'yellow': 'Amarelo',
    'green': 'Verde'
}

# Crie um selectbox para o usuário selecionar a cor
cor_selecionada = st.selectbox('Selecione a cor para filtrar', list(cores.values()))

# Use a função verificar_cor para filtrar o DataFrame

if cor_selecionada == 'Todas':
    styled_filtrado = df.style.apply(estilizar_colunas, axis=1)
else:
    filtrado = df[df.apply(verificar_cor, cor=list(cores.keys())[list(cores.values()).index(cor_selecionada)], axis=1)]
    styled_filtrado = filtrado.style.apply(aplicar_estilo_cor, cor=list(cores.keys())[list(cores.values()).index(cor_selecionada)], axis=1)

# Exiba o DataFrame filtrado e estilizado
st.dataframe(styled_filtrado)

# Trazer o nome do arquivo digitado pelo usuário
file_name = st.text_input('Digite o nome do arquivo para salvar:')

# Botão de exportação
if st.button('Exportar para Excel'):
    if file_name:
        with pd.ExcelWriter(f'{file_name}.xlsx', engine='openpyxl') as writer:
            styled_filtrado.to_excel(writer, sheet_name='Volumes por Unidades de Negócio', index=False, startrow=1)
            
            worksheet = writer.sheets['Volumes por Unidades de Negócio']
            
            # Adicionar imagem no canto superior esquerdo da planilha
            img_path = 'images/logo.png'
            img = openpyxl.drawing.image.Image(img_path)
            worksheet.add_image(img, 'A1')  # Ajustar a referência da célula conforme necessário
            
            # Adicionar título ao lado da imagem
            title_cell = worksheet.cell(row=1, column=7)  # Célula G1
            title_cell.value = 'Análise de Volumes Por Unidade de Negócio'
            title_cell.font = openpyxl.styles.Font(size=12, bold=True, name='Arial', color='00808080')  # Definir cor da fonte como #808080
            title_cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
            
            # Adicionar data e hora de geração
            current_datetime = datetime.now().strftime("%d/%m/%Y %H:%M")
            generated_cell = worksheet.cell(row=1, column=8)  # Célula G1
            generated_cell.value = f'Gerado em: {current_datetime}'
            generated_cell.font = openpyxl.styles.Font(size=12, bold=True, name='Arial', color='00808080')
            generated_cell.alignment = Alignment(wrap_text=True)
            
            legenda = [
                ["", "Valores igual a 0"],
                ["", "Valores duplicados"],
                ["", "Diferença porcentagem >= 50%"],
                ["", "Diferença porcentagem >= 30%"],
                ["", "Diferença porcentagem >= 15%"],
                ["", "Diferença porcentagem >= 5%"]
            ]

            colors = [
                "FF90EE90",  # verde
                "FF87CEEB",  # azul claro
                "FFFF0000",  # vermelho
                "FFFFA500",  # laranja
                "FFFFFF00",  # amarelo
                "FF00FF00"  # verde
            ]

            for row, (row_data, color) in enumerate(zip(legenda, colors), start=2):  # Iniciar a partir da linha 4
                for col, value in enumerate(row_data, start=10):  # Iniciar a partir da coluna 10 (J)
                    cell = worksheet.cell(row=row, column=col)
                    cell.value = value
                    if col == 10:  # Preencher apenas a segunda coluna (K)
                        cell.fill = openpyxl.styles.PatternFill(start_color=color, fill_type='solid')

            st.success(f'Arquivo {file_name}.xlsx exportado com sucesso!')
    else:
        st.error('Por favor, digite um nome para o arquivo.')